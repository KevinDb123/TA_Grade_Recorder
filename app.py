from __future__ import annotations

import os
import re
import stat
import sys
import warnings
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from shutil import copyfileobj
from threading import RLock
from typing import Any

from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename


if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
    APP_DIR = Path(sys._MEIPASS)
else:
    APP_DIR = Path(__file__).resolve().parent

DATA_DIR = Path(os.environ.get("TA_GRADE_RECORDER_DATA_DIR", APP_DIR)).resolve()
HEADER_ROW = 6
SUBHEADER_ROW = 7
DATA_START_ROW = 8
STUDENT_ID_COL = 2
STUDENT_NAME_COL = 3
GRADE_SECTION_TITLE = "平时成绩"
VALID_BULK_MODES = {"overwrite", "skip_existing"}

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
)


@dataclass
class Student:
    row: int
    student_id: str
    name: str
    scores: dict[str, Any]


@dataclass
class GradeSection:
    start_col: int
    end_col: int


class GradeBook:
    def __init__(self, base_dir: Path) -> None:
        self.base_dir = base_dir
        self.lock = RLock()
        self.workbook_path = self._find_workbook()

    def list_workbooks(self) -> list[Path]:
        return sorted(self.base_dir.glob("*.xlsx"))

    def workbook_choices(self) -> list[dict[str, Any]]:
        return [
            {"name": file.name, "selected": file == self.workbook_path}
            for file in self.list_workbooks()
        ]

    def _find_workbook(self) -> Path | None:
        files = self.list_workbooks()
        return files[0] if files else None

    def _require_workbook(self) -> Path:
        if self.workbook_path is None:
            raise ValueError("请先导入或选择 Excel 文件。")
        return self.workbook_path

    def set_workbook(self, filename: str) -> dict[str, str]:
        target = next((file for file in self.list_workbooks() if file.name == filename), None)
        if target is None:
            raise ValueError("未找到指定的 xlsx 文件。")
        with self.lock:
            self.workbook_path = target
        return {"name": target.name}

    def import_workbook(self, file_storage) -> dict[str, str]:
        original_name = file_storage.filename or ""
        suffix = Path(original_name).suffix.lower()
        if suffix != ".xlsx":
            raise ValueError("只能导入 xlsx 文件。")

        base_name = secure_filename(Path(original_name).stem) or "uploaded_workbook"
        candidate = self.base_dir / f"{base_name}.xlsx"
        counter = 1
        while candidate.exists():
            candidate = self.base_dir / f"{base_name}_{counter}.xlsx"
            counter += 1

        file_storage.stream.seek(0)
        with candidate.open("wb") as handle:
            copyfileobj(file_storage.stream, handle)

        with self.lock:
            self.workbook_path = candidate
        return {"name": candidate.name}

    def _load_workbook(self):
        workbook = load_workbook(self._require_workbook())
        sheet = workbook[workbook.sheetnames[0]]
        return workbook, sheet

    def _save_workbook(self, workbook) -> None:
        try:
            workbook_path = self._require_workbook()
            workbook_path.chmod(workbook_path.stat().st_mode | stat.S_IWRITE)
            workbook.save(workbook_path)
        except PermissionError as exc:
            raise PermissionError(
                "当前 Excel 文件无法写入，请先关闭占用该文件的 Excel 或 WPS 后再重试。"
            ) from exc

    def _find_grade_section(self, sheet) -> GradeSection:
        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(HEADER_ROW, col).value
            if str(value).strip() != GRADE_SECTION_TITLE:
                continue

            for merged_range in sheet.merged_cells.ranges:
                if (
                    merged_range.min_row <= HEADER_ROW <= merged_range.max_row
                    and merged_range.min_col <= col <= merged_range.max_col
                ):
                    return GradeSection(
                        start_col=merged_range.min_col,
                        end_col=merged_range.max_col,
                    )

            end_col = col
            while end_col + 1 <= sheet.max_column and not sheet.cell(HEADER_ROW, end_col + 1).value:
                end_col += 1
            return GradeSection(start_col=col, end_col=end_col)

        raise ValueError("未找到“平时成绩”所在的表头区域。")

    def _default_experiment_name(self, position: int) -> str:
        return f"第{position}次"

    def _display_score(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        numeric = float(value)
        return str(int(numeric)) if numeric.is_integer() else str(numeric)

    def _students_from_sheet(self, sheet, experiments: list[dict[str, str]]) -> list[Student]:
        students: list[Student] = []
        row = DATA_START_ROW
        while True:
            student_id = sheet.cell(row, STUDENT_ID_COL).value
            name = sheet.cell(row, STUDENT_NAME_COL).value
            if not student_id and not name:
                break
            scores = {
                item["column"]: sheet.cell(row, self._column_index(item["column"])).value
                for item in experiments
            }
            students.append(
                Student(
                    row=row,
                    student_id=str(student_id or "").strip(),
                    name=str(name or "").strip(),
                    scores=scores,
                )
            )
            row += 1
        return students

    def _serialize_experiments_from_sheet(self, sheet) -> list[dict[str, str]]:
        section = self._find_grade_section(sheet)
        experiments: list[dict[str, str]] = []
        for position, col in enumerate(range(section.start_col, section.end_col + 1), start=1):
            raw_name = sheet.cell(SUBHEADER_ROW, col).value
            name = str(raw_name).strip() if raw_name is not None and str(raw_name).strip() else self._default_experiment_name(position)
            experiments.append(
                {
                    "key": get_column_letter(col),
                    "name": name,
                    "column": get_column_letter(col),
                }
            )
        return experiments

    def get_experiments(self) -> list[dict[str, str]]:
        if self.workbook_path is None:
            return []
        with self.lock:
            workbook, sheet = self._load_workbook()
            try:
                return self._serialize_experiments_from_sheet(sheet)
            finally:
                workbook.close()

    def get_preview_info(self) -> dict[str, Any]:
        if self.workbook_path is None:
            return {
                "workbook_name": "",
                "grade_range": "",
                "student_id_column": get_column_letter(STUDENT_ID_COL),
                "student_name_column": get_column_letter(STUDENT_NAME_COL),
                "data_start_row": DATA_START_ROW,
            }
        with self.lock:
            workbook, sheet = self._load_workbook()
            try:
                section = self._find_grade_section(sheet)
                return {
                    "workbook_name": self.workbook_path.name,
                    "grade_range": f"{get_column_letter(section.start_col)}:{get_column_letter(section.end_col)}",
                    "student_id_column": get_column_letter(STUDENT_ID_COL),
                    "student_name_column": get_column_letter(STUDENT_NAME_COL),
                    "data_start_row": DATA_START_ROW,
                }
            finally:
                workbook.close()

    def list_students(self) -> list[Student]:
        with self.lock:
            workbook, sheet = self._load_workbook()
            try:
                experiments = self._serialize_experiments_from_sheet(sheet)
                return self._students_from_sheet(sheet, experiments)
            finally:
                workbook.close()

    def search_students(self, keyword: str) -> list[dict[str, Any]]:
        normalized = keyword.strip().lower()
        students = self.list_students()
        if normalized:
            students = [
                item
                for item in students
                if normalized in item.student_id.lower() or normalized in item.name.lower()
            ]
        return [
            {
                "row": item.row,
                "student_id": item.student_id,
                "name": item.name,
                "scores": item.scores,
            }
            for item in students
        ]

    def report_for_experiment(self, column_key: str) -> dict[str, Any]:
        column = self._column_from_key(column_key)
        letter = get_column_letter(column)
        experiments = self.get_experiments()
        experiment = next((item for item in experiments if item["column"] == letter), None)
        if experiment is None:
            raise ValueError("当前次数列不存在。")

        rows = []
        scores: list[float] = []
        for student in self.list_students():
            raw_score = student.scores.get(letter)
            score = None if raw_score in (None, "") else float(raw_score)
            if score is not None:
                scores.append(score)
            rows.append(
                {
                    "student_id": student.student_id,
                    "name": student.name,
                    "display_name": student.name[:-1] if student.name.endswith("*") else student.name,
                    "gender": "女" if student.name.endswith("*") else "男",
                    "score": score,
                    "score_text": self._display_score(raw_score),
                }
            )

        rows.sort(key=lambda item: (item["score"] is None, -(item["score"] or 0), item["student_id"]))

        total_students = len(rows)
        scored_students = len(scores)
        missing_students = total_students - scored_students
        average_score = round(sum(scores) / scored_students, 2) if scored_students else None

        return {
            "experiment": experiment,
            "summary": {
                "total_students": total_students,
                "scored_students": scored_students,
                "missing_students": missing_students,
                "average_score": average_score,
                "max_score": max(scores) if scores else None,
                "min_score": min(scores) if scores else None,
            },
            "rows": rows,
        }

    def rename_experiment(self, column_key: str, name: str) -> dict[str, Any]:
        column = self._column_from_key(column_key)
        cleaned_name = name.strip()
        if not cleaned_name:
            raise ValueError("次数名称不能为空。")
        with self.lock:
            workbook, sheet = self._load_workbook()
            try:
                sheet.cell(SUBHEADER_ROW, column).value = cleaned_name
                self._save_workbook(workbook)
            finally:
                workbook.close()
        return {
            "column": get_column_letter(column),
            "name": cleaned_name,
        }

    def add_experiment(self, name: str | None = None) -> dict[str, Any]:
        with self.lock:
            workbook, sheet = self._load_workbook()
            try:
                section = self._find_grade_section(sheet)
                experiments = self._serialize_experiments_from_sheet(sheet)
                new_name = name.strip() if name and name.strip() else self._default_experiment_name(len(experiments) + 1)
                insert_at = section.end_col + 1
                self._modify_columns_with_merged_cells(sheet, "insert", insert_at, 1)
                self._copy_column_style(sheet, insert_at - 1, insert_at)
                self._reset_grade_header_merge(sheet, section.start_col, section.end_col + 1)
                sheet.cell(SUBHEADER_ROW, insert_at).value = new_name
                self._save_workbook(workbook)
                refreshed_experiments = self._serialize_experiments_from_sheet(sheet)
            finally:
                workbook.close()
        return {
            "message": "已新增平时成绩列。",
            "experiment": {
                "column": get_column_letter(insert_at),
                "name": new_name,
            },
            "experiments": refreshed_experiments,
        }

    def delete_experiment(self, column_key: str) -> dict[str, Any]:
        column = self._column_from_key(column_key)
        with self.lock:
            workbook, sheet = self._load_workbook()
            try:
                section = self._find_grade_section(sheet)
                if section.end_col - section.start_col < 1:
                    raise ValueError("至少需要保留 1 列平时成绩。")
                self._modify_columns_with_merged_cells(sheet, "delete", column, 1)
                self._save_workbook(workbook)
                refreshed_experiments = self._serialize_experiments_from_sheet(sheet)
            finally:
                workbook.close()
        return {
            "message": "已删除平时成绩列。",
            "experiments": refreshed_experiments,
        }

    def update_score(self, column_key: str, student_row: int, score: float | int | str) -> dict[str, Any]:
        column = self._column_from_key(column_key)
        parsed_score = self._parse_score(score)
        with self.lock:
            workbook, sheet = self._load_workbook()
            try:
                if student_row < DATA_START_ROW:
                    raise ValueError("学生行号不合法。")
                sheet.cell(student_row, column).value = parsed_score
                self._save_workbook(workbook)
            finally:
                workbook.close()
        return {}

    def preview_bulk_update(
        self,
        column_key: str,
        entries_text: str,
        same_score: float | int | str | None = None,
        mode: str = "overwrite",
    ) -> dict[str, Any]:
        if mode not in VALID_BULK_MODES:
            raise ValueError("批量录分模式不支持。")

        column = self._column_from_key(column_key)
        letter = get_column_letter(column)
        entries = self._parse_bulk_entries(entries_text, same_score)
        students = self.list_students()
        student_map = {item.student_id.lower(): item for item in students}
        student_name_map = {item.name.lower(): item for item in students}

        rows: list[dict[str, Any]] = []
        summary = {
            "total_lines": len(entries),
            "matched_count": 0,
            "missing_count": 0,
            "will_update_count": 0,
            "skip_existing_count": 0,
            "overwrite_count": 0,
            "fill_empty_count": 0,
        }

        for key, score in entries:
            target = student_map.get(key.lower()) or student_name_map.get(key.lower())
            if not target:
                summary["missing_count"] += 1
                rows.append(
                    {
                        "input": key,
                        "matched": False,
                        "status": "missing",
                        "status_text": "未匹配",
                        "student_id": "",
                        "name": "",
                        "current_score_text": "",
                        "new_score_text": self._display_score(score),
                    }
                )
                continue

            summary["matched_count"] += 1
            current_score = target.scores.get(letter)
            has_existing_score = current_score not in (None, "")

            if has_existing_score and mode == "skip_existing":
                summary["skip_existing_count"] += 1
                status = "skipped"
                status_text = "已跳过"
            else:
                summary["will_update_count"] += 1
                if has_existing_score:
                    summary["overwrite_count"] += 1
                    status = "overwrite"
                    status_text = "将覆盖"
                else:
                    summary["fill_empty_count"] += 1
                    status = "fill"
                    status_text = "将写入"

            rows.append(
                {
                    "input": key,
                    "matched": True,
                    "status": status,
                    "status_text": status_text,
                    "student_id": target.student_id,
                    "name": target.name,
                    "current_score_text": self._display_score(current_score),
                    "new_score_text": self._display_score(score),
                }
            )

        return {"summary": summary, "rows": rows}

    def bulk_update_scores(
        self,
        column_key: str,
        entries_text: str,
        same_score: float | int | str | None = None,
        mode: str = "overwrite",
    ) -> dict[str, Any]:
        preview = self.preview_bulk_update(column_key, entries_text, same_score, mode)
        if preview["summary"]["matched_count"] == 0:
            raise ValueError("没有可写入的匹配学生。")

        column = self._column_from_key(column_key)
        entries = self._parse_bulk_entries(entries_text, same_score)
        students = self.list_students()
        student_map = {item.student_id.lower(): item for item in students}
        student_name_map = {item.name.lower(): item for item in students}

        updated: list[str] = []
        missing: list[str] = []
        skipped: list[str] = []

        with self.lock:
            workbook, sheet = self._load_workbook()
            try:
                for key, score in entries:
                    target = student_map.get(key.lower()) or student_name_map.get(key.lower())
                    if not target:
                        missing.append(key)
                        continue

                    current_value = sheet.cell(target.row, column).value
                    if current_value not in (None, "") and mode == "skip_existing":
                        skipped.append(f"{target.student_id} {target.name}")
                        continue

                    sheet.cell(target.row, column).value = self._parse_score(score)
                    updated.append(f"{target.student_id} {target.name}")

                self._save_workbook(workbook)
            finally:
                workbook.close()

        return {
            "updated": updated,
            "missing": missing,
            "skipped": skipped,
            "count": len(updated),
            "preview": preview,
        }

    def _parse_bulk_entries(
        self,
        entries_text: str,
        same_score: float | int | str | None,
    ) -> list[tuple[str, float]]:
        lines = [line.strip() for line in entries_text.splitlines() if line.strip()]
        if not lines:
            raise ValueError("请至少输入一条批量成绩记录。")

        parsed_entries: list[tuple[str, float]] = []
        if same_score not in (None, ""):
            shared_score = self._parse_score(same_score)
            for line in lines:
                parsed_entries.append((line, shared_score))
            return parsed_entries

        for line in lines:
            parts = re.split(r"[\s,，]+", line)
            if len(parts) < 2:
                raise ValueError(f"无法识别这一行：{line}")
            parsed_entries.append((parts[0], self._parse_score(parts[-1])))
        return parsed_entries

    def _column_from_key(self, column_key: str) -> int:
        key = column_key.strip().upper()
        for item in self.get_experiments():
            if item["column"] == key:
                return self._column_index(item["column"])
        raise ValueError("当前次数列不存在。")

    def _column_index(self, column_key: str) -> int:
        value = 0
        for char in column_key:
            value = value * 26 + (ord(char) - ord("A") + 1)
        return value

    def _parse_score(self, score: float | int | str) -> float:
        text = str(score).strip()
        if text == "":
            raise ValueError("成绩不能为空。")
        value = float(text)
        if value < 0:
            raise ValueError("成绩不能小于 0。")
        return int(value) if value.is_integer() else value

    def _modify_columns_with_merged_cells(self, sheet, action: str, index: int, amount: int) -> None:
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))

        if action == "insert":
            sheet.insert_cols(index, amount)
            transformed_ranges = [
                self._transform_merged_range_for_insert(merged_range, index, amount)
                for merged_range in merged_ranges
            ]
        elif action == "delete":
            sheet.delete_cols(index, amount)
            transformed_ranges = [
                self._transform_merged_range_for_delete(merged_range, index, amount)
                for merged_range in merged_ranges
            ]
        else:
            raise ValueError("不支持的列操作。")

        for merged_range in transformed_ranges:
            if merged_range is None:
                continue
            min_row, min_col, max_row, max_col = merged_range
            if min_row == max_row and min_col == max_col:
                continue
            sheet.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )

    def _reset_grade_header_merge(self, sheet, start_col: int, end_col: int) -> None:
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            if (
                merged_range.min_row <= HEADER_ROW <= merged_range.max_row
                and merged_range.min_col == start_col
                and str(sheet.cell(HEADER_ROW, start_col).value).strip() == GRADE_SECTION_TITLE
            ):
                sheet.unmerge_cells(str(merged_range))
        sheet.cell(HEADER_ROW, start_col).value = GRADE_SECTION_TITLE
        if end_col > start_col:
            sheet.merge_cells(
                start_row=HEADER_ROW,
                start_column=start_col,
                end_row=HEADER_ROW,
                end_column=end_col,
            )

    def _copy_column_style(self, sheet, source_col: int, target_col: int) -> None:
        for row in range(1, sheet.max_row + 1):
            source_cell = sheet.cell(row, source_col)
            target_cell = sheet.cell(row, target_col)
            if isinstance(target_cell, MergedCell):
                continue
            target_cell.value = None
            target_cell._style = copy(source_cell._style)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)

        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(target_col)
        if source_letter in sheet.column_dimensions:
            sheet.column_dimensions[target_letter].width = sheet.column_dimensions[source_letter].width
            sheet.column_dimensions[target_letter].hidden = sheet.column_dimensions[source_letter].hidden
            sheet.column_dimensions[target_letter].bestFit = sheet.column_dimensions[source_letter].bestFit

    def _transform_merged_range_for_insert(self, merged_range, index: int, amount: int):
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        if max_col < index:
            return min_row, min_col, max_row, max_col
        if min_col >= index:
            return min_row, min_col + amount, max_row, max_col + amount
        if min_col < index <= max_col:
            return min_row, min_col, max_row, max_col + amount
        return min_row, min_col, max_row, max_col

    def _transform_merged_range_for_delete(self, merged_range, index: int, amount: int):
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        delete_start = index
        delete_end = index + amount - 1

        if max_col < delete_start:
            return min_row, min_col, max_row, max_col
        if min_col > delete_end:
            return min_row, min_col - amount, max_row, max_col - amount
        if min_col >= delete_start and max_col <= delete_end:
            return None
        if min_col < delete_start and max_col > delete_end:
            return min_row, min_col, max_row, max_col - amount
        if min_col < delete_start <= max_col <= delete_end:
            return min_row, min_col, max_row, delete_start - 1
        if delete_start <= min_col <= delete_end < max_col:
            new_width = max_col - delete_end
            return min_row, delete_start, max_row, delete_start + new_width - 1
        return min_row, min_col, max_row, max_col


app = Flask(
    __name__,
    template_folder=str(APP_DIR / "templates"),
    static_folder=str(APP_DIR / "static"),
)
gradebook = GradeBook(DATA_DIR)


@app.get("/")
def index():
    return render_template(
        "index.html",
        workbook_name=gradebook.workbook_path.name if gradebook.workbook_path else "未选择文件",
        workbook_choices=gradebook.workbook_choices(),
        experiments=gradebook.get_experiments(),
    )


@app.get("/report")
def report_page():
    experiments = gradebook.get_experiments()
    default_column = experiments[0]["column"] if experiments else ""
    requested_column = request.args.get("column", default_column)
    current_column = requested_column or default_column
    if current_column:
        try:
            report = gradebook.report_for_experiment(current_column)
        except ValueError:
            current_column = default_column
            report = gradebook.report_for_experiment(current_column) if current_column else {
                "experiment": {"name": "未选择次数", "column": ""},
                "summary": {
                    "total_students": 0,
                    "scored_students": 0,
                    "missing_students": 0,
                    "average_score": None,
                    "max_score": None,
                    "min_score": None,
                },
                "rows": [],
            }
    else:
        report = {
            "experiment": {"name": "未选择次数", "column": ""},
            "summary": {
                "total_students": 0,
                "scored_students": 0,
                "missing_students": 0,
                "average_score": None,
                "max_score": None,
                "min_score": None,
            },
            "rows": [],
        }
    return render_template(
        "report.html",
        workbook_name=gradebook.workbook_path.name if gradebook.workbook_path else "未选择文件",
        workbook_choices=gradebook.workbook_choices(),
        experiments=experiments,
        current_column=current_column,
        report=report,
    )


@app.get("/api/students")
def api_students():
    keyword = request.args.get("q", "")
    return jsonify({"students": gradebook.search_students(keyword)})


@app.get("/api/report")
def api_report():
    column = request.args.get("column", "")
    return jsonify(gradebook.report_for_experiment(column))


@app.get("/health")
def health_check():
    return jsonify({"status": "ok"})


@app.post("/api/workbooks/select")
def api_select_workbook():
    payload = request.get_json(force=True)
    result = gradebook.set_workbook(payload["filename"])
    return jsonify(
        {
            "message": "已切换 Excel 文件。",
            "workbook": result,
            "workbooks": gradebook.workbook_choices(),
            "experiments": gradebook.get_experiments(),
        }
    )


@app.post("/api/workbooks/upload")
def api_upload_workbook():
    file = request.files.get("file")
    if file is None or not file.filename:
        raise ValueError("请选择要导入的 xlsx 文件。")
    result = gradebook.import_workbook(file)
    return jsonify(
        {
            "message": "已成功导入 Excel 文件。",
            "workbook": result,
            "workbooks": gradebook.workbook_choices(),
            "experiments": gradebook.get_experiments(),
        }
    )


@app.post("/api/experiments/add")
def api_add_experiment():
    payload = request.get_json(silent=True) or {}
    return jsonify(gradebook.add_experiment(payload.get("name")))


@app.post("/api/experiments/rename")
def api_rename_experiment():
    payload = request.get_json(force=True)
    result = gradebook.rename_experiment(payload["column"], payload["name"])
    return jsonify({"message": "当前次数名称已保存。", "experiment": result})


@app.post("/api/experiments/delete")
def api_delete_experiment():
    payload = request.get_json(force=True)
    return jsonify(gradebook.delete_experiment(payload["column"]))


@app.post("/api/scores/single")
def api_update_single():
    payload = request.get_json(force=True)
    result = gradebook.update_score(payload["column"], int(payload["row"]), payload["score"])
    return jsonify({"message": "成绩已写入 Excel。", **result})


@app.post("/api/scores/bulk/preview")
def api_bulk_preview():
    payload = request.get_json(force=True)
    result = gradebook.preview_bulk_update(
        payload["column"],
        payload.get("entries", ""),
        payload.get("same_score"),
        payload.get("mode", "overwrite"),
    )
    return jsonify(result)


@app.post("/api/scores/bulk")
def api_update_bulk():
    payload = request.get_json(force=True)
    result = gradebook.bulk_update_scores(
        payload["column"],
        payload.get("entries", ""),
        payload.get("same_score"),
        payload.get("mode", "overwrite"),
    )
    return jsonify({"message": f"已更新 {result['count']} 名学生。", **result})


@app.errorhandler(ValueError)
def handle_value_error(error):
    return jsonify({"message": str(error)}), 400


@app.errorhandler(PermissionError)
def handle_permission_error(error):
    return jsonify({"message": str(error)}), 423


@app.errorhandler(Exception)
def handle_error(error):
    status_code = getattr(error, "code", 500)
    return jsonify({"message": str(error)}), status_code


if __name__ == "__main__":
    app.run(
        host=os.environ.get("TA_GRADE_RECORDER_HOST", "127.0.0.1"),
        port=int(os.environ.get("TA_GRADE_RECORDER_PORT", "5000")),
        debug=False,
        use_reloader=False,
    )
